"""HTTP scraping adapter for generic web URLs.

Migrated and normalized from ``lmforge_core.views.scrape``.
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from io import BytesIO
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .content_extractor import extract_article_content

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class GenericWebResult:
    """Normalized payload returned by generic URL parsing."""

    file_type: str
    content: str
    extracted_title: str = ""


class GenericWebScraper:
    """Scraper for non-Reddit/non-PDF URL content types."""

    def scrape(self, url: str, timeout: int = 30) -> GenericWebResult:
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()

        content_type = response.headers.get("content-type", "").lower()

        if "application/json" in content_type or (response.text and response.text.strip().startswith("{")):
            return GenericWebResult(file_type="json", content=json.dumps(response.json(), indent=2))

        if any(x in content_type for x in ("application/xml", "text/xml", "application/rss+xml")):
            return GenericWebResult(file_type="xml", content=response.text)

        if "text/plain" in content_type:
            return GenericWebResult(file_type="text", content=response.text)

        if "text/csv" in content_type or url.lower().endswith(".csv"):
            return GenericWebResult(file_type="csv", content=response.text)

        if any(x in content_type for x in ("excel", "spreadsheetml", "vnd.openxmlformats")) or url.lower().endswith(".xlsx"):
            return GenericWebResult(file_type="xlsx", content=self._parse_xlsx_text(response.content))

        # Default: HTML/web article extraction
        return self._parse_html(response.content, url=url)

    def _parse_xlsx_text(self, raw_bytes: bytes) -> str:
        bio = BytesIO(raw_bytes)
        wb = load_workbook(filename=bio, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [",".join([str(c) if c is not None else "" for c in row]) for row in ws.iter_rows(values_only=True)]
        return "\n".join(rows)

    def _parse_html(self, raw_html: bytes, *, url: str) -> GenericWebResult:
        extracted_title = ""
        try:
            result: dict[str, Any] = extract_article_content(raw_html, url)
            content = (result.get("body") or "").strip()
            extracted_title = str(result.get("title") or "")

            if content:
                return GenericWebResult(file_type="html", content=content, extracted_title=extracted_title)
        except Exception as exc:
            logger.exception("Article extractor failed for %s: %s", url, exc)

        # safe fallback
        soup = BeautifulSoup(raw_html, "html.parser")
        article = soup.find("article") or soup.find(class_="content") or soup.find("main")
        content = (article or soup).get_text("\n\n", strip=True)
        if not extracted_title and soup.title and soup.title.string:
            extracted_title = soup.title.string.strip()

        return GenericWebResult(file_type="html", content=content, extracted_title=extracted_title)
