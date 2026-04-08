"""Application service: source ingestion and chunk preparation."""

from __future__ import annotations

from dataclasses import dataclass
import json
import re
from io import BytesIO
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from studio.domain.models.source_documents import SourceDocument
from studio.infrastructure.scraping.content_extractor import extract_article_content


EMOJI_RE = re.compile(r"[\U00010000-\U0010FFFF]")


@dataclass(slots=True)
class ScrapedPayload:
    url: str
    file_type: str
    title: str
    content: str


class DocumentService:
    """Service migrated from `lmforge_core.views.scrape` and dataset chunking flows."""

    max_title_length = 100
    max_url_title_length = 95

    def remove_emojis(self, text: str) -> str:
        return EMOJI_RE.sub("", text or "")

    def split_text(self, text: str, max_tokens: int = 1000, tokenizer: Any | None = None) -> list[str]:
        """Paragraph-based chunking compatible with legacy workflow behavior."""
        paragraphs = (text or "").split("\n\n")
        chunks: list[str] = []
        current_chunk: list[str] = []
        current_tokens = 0

        for paragraph in paragraphs:
            para_tokens = len(tokenizer.encode(paragraph)) if tokenizer else len(paragraph.split())
            if current_tokens + para_tokens > max_tokens and current_chunk:
                chunks.append("\n\n".join(current_chunk))
                current_chunk = [paragraph]
                current_tokens = para_tokens
            else:
                current_chunk.append(paragraph)
                current_tokens += para_tokens

        if current_chunk:
            chunks.append("\n\n".join(current_chunk))
        return [chunk for chunk in chunks if chunk.strip()]

    def scrape_generic_url(self, url: str, title: str = "") -> ScrapedPayload:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        content_type = response.headers.get("content-type", "").lower()

        content = ""
        file_type = "text"

        if "application/json" in content_type or (response.text and response.text.strip().startswith("{")):
            content = json.dumps(response.json(), indent=2)
            file_type = "json"
        elif any(x in content_type for x in ("application/xml", "text/xml", "application/rss+xml")):
            content = response.text
            file_type = "xml"
        elif "text/plain" in content_type:
            content = response.text
            file_type = "text"
        elif "text/csv" in content_type or url.lower().endswith(".csv"):
            content = response.text
            file_type = "csv"
        elif any(x in content_type for x in ("excel", "spreadsheetml", "vnd.openxmlformats")) or url.lower().endswith(".xlsx"):
            bio = BytesIO(response.content)
            wb = load_workbook(filename=bio, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [",".join([str(c) if c is not None else "" for c in row]) for row in ws.iter_rows(values_only=True)]
            content = "\n".join(rows)
            file_type = "xlsx"
        else:
            result: dict[str, Any] = extract_article_content(response.content, url)
            content = (result.get("body") or "").strip()
            if not content:
                soup = BeautifulSoup(response.content, "html.parser")
                article = soup.find("article") or soup.find(class_="content") or soup.find("main")
                content = (article or soup).get_text("\n\n", strip=True)
            if not title.strip() and result.get("title"):
                title = str(result["title"])
            file_type = "html"

        final_title = (title or url[: self.max_url_title_length] or "scraped")[: self.max_title_length]
        return ScrapedPayload(url=url, file_type=file_type, title=final_title, content=self.remove_emojis(content))

    def persist_source_document(self, payload: ScrapedPayload) -> SourceDocument:
        return SourceDocument.objects.create(
            url=payload.url,
            file_type=payload.file_type,
            content=payload.content,
            title=payload.title,
        )
